"""
Excel 出力モジュール（openpyxl）
WBS / 直近のタスク / 進捗一覧 の 3 シートを生成する
"""
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _fmt_date(d):
    """日付を YYYY/MM/DD 形式の文字列に変換する"""
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%Y/%m/%d')
    return str(d)


def _apply_header_style(ws, row_no, col_count):
    """ヘッダー行にダークブルーの背景・白太字を適用する"""
    fill = PatternFill(fill_type='solid', fgColor='366092')
    font = Font(bold=True, size=9, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_no, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _set_col_widths(ws, widths):
    """列幅を一括設定する"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze_header(ws):
    """1行目をフリーズする"""
    ws.freeze_panes = 'A2'


# ============================================================
# タスクツリー構築ユーティリティ
# ============================================================

def _load_tasks(project):
    """プロジェクトの全タスク（論理削除済み除外）をロードする"""
    from apps.tasks.models import Task
    return list(
        Task.objects.filter(
            project=project,
            deleted_at__isnull=True,
        ).select_related('quarter').prefetch_related('assignees__user')
        .order_by('wbs_no')
    )


def _build_children_map(all_tasks):
    """parent_id → 子リストのマップを構築する"""
    children_map: dict = {}
    roots = []
    for task in all_tasks:
        if task.parent_task_id is None:
            roots.append(task)
        else:
            children_map.setdefault(task.parent_task_id, []).append(task)
    roots.sort(key=lambda t: t.order)
    for lst in children_map.values():
        lst.sort(key=lambda t: t.order)
    return roots, children_map


def _calc_visible_ids(all_tasks, quarter_id=None, start_date=None, end_date=None):
    """フィルタに一致するタスク ID + その祖先 ID を返す。フィルタなしは None を返す。"""
    if not quarter_id and not (start_date and end_date):
        return None  # フィルタなし = 全件

    matched = set()
    for task in all_tasks:
        if quarter_id:
            if task.quarter_id and str(task.quarter_id) == str(quarter_id):
                matched.add(task.id)
        else:
            # 日付オーバーラップ判定
            if task.start_date and task.end_date:
                if task.start_date <= end_date and task.end_date >= start_date:
                    matched.add(task.id)

    # 祖先 ID を追加してパンくず行が表示されるようにする
    id_map = {t.id: t for t in all_tasks}
    ancestors: set = set()
    for tid in list(matched):
        node = id_map.get(tid)
        while node and node.parent_task_id:
            ancestors.add(node.parent_task_id)
            node = id_map.get(node.parent_task_id)

    return matched | ancestors


def _walk_tree(roots, children_map, visible_ids=None, max_depth=3):
    """深さ優先でツリーを走査してフラットリストを返す"""
    result = []

    def _walk(task, depth):
        if visible_ids is not None and task.id not in visible_ids:
            return
        if depth > max_depth:
            return
        result.append(task)
        for child in children_map.get(task.id, []):
            _walk(child, depth + 1)

    for root in roots:
        _walk(root, 0)
    return result


def _collect_actual_tasks(task, children_map):
    """ノード配下の実タスク（task_type='task'）を再帰収集する"""
    result = []
    if task.task_type == 'task':
        result.append(task)
    for child in children_map.get(task.id, []):
        result.extend(_collect_actual_tasks(child, children_map))
    return result


# ============================================================
# Sheet 1: WBS
# ============================================================

_DEPTH_FILLS = ['EFF3FF', 'F5F5F5', 'FFFFFF', 'FAFAFA']
_INDENT = ['', '  ', '    ', '      ']

def _build_wbs_sheet(ws, project, quarter_id, start_date, end_date):
    all_tasks = _load_tasks(project)
    visible_ids = _calc_visible_ids(all_tasks, quarter_id, start_date, end_date)
    roots, children_map = _build_children_map(all_tasks)
    tasks = _walk_tree(roots, children_map, visible_ids, max_depth=3)

    headers = [
        'WBS No', 'タスク名', '種別', 'ステータス',
        '担当者', 'クォーター',
        '開始日（予定）', '終了日（予定）', '開始日（実績）', '終了日（実績）',
        '見積(h)', '進捗%',
    ]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers))
    _freeze_header(ws)

    for task in tasks:
        assignees = ', '.join(a.user.username for a in task.assignees.all() if a.user)
        quarter_title = task.quarter.title if task.quarter else ''
        estimated = float(task.estimated_hours) if task.estimated_hours else ''
        progress = f'{task.progress}%'
        depth = min(task.depth, 3)

        ws.append([
            task.wbs_no,
            _INDENT[depth] + task.title,
            task.task_type,
            task.status,
            assignees,
            quarter_title,
            _fmt_date(task.start_date),
            _fmt_date(task.end_date),
            _fmt_date(task.actual_start_date),
            _fmt_date(task.actual_end_date),
            estimated,
            progress,
        ])

        fill = PatternFill(fill_type='solid', fgColor=_DEPTH_FILLS[depth])
        font = Font(size=9, bold=(task.depth == 0))
        align = Alignment(vertical='center')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = align

    _set_col_widths(ws, [10, 40, 8, 12, 18, 14, 14, 14, 14, 14, 8, 8])


# ============================================================
# Sheet 2: 直近のタスク
# ============================================================

def _build_recent_sheet(ws, project, start_date, end_date):
    from apps.tasks.models import Task

    today = date.today()
    week_later = today + timedelta(days=7)

    base_qs = Task.objects.filter(
        project=project,
        deleted_at__isnull=True,
        start_date__isnull=False,
        end_date__isnull=False,
        actual_end_date__isnull=True,
    ).exclude(status='Done').prefetch_related('assignees__user')

    # 期間フィルタ（指定があれば）
    if start_date and end_date:
        base_qs = base_qs.filter(start_date__lte=end_date, end_date__gte=start_date)

    overdue = list(base_qs.filter(end_date__lt=today).order_by('end_date'))
    used_ids = {t.id for t in overdue}

    starting_soon = list(
        base_qs.filter(start_date__gte=today, start_date__lte=week_later)
        .exclude(id__in=used_ids).order_by('start_date')
    )
    used_ids |= {t.id for t in starting_soon}

    in_progress = list(
        base_qs.filter(actual_start_date__isnull=False)
        .exclude(id__in=used_ids).order_by('actual_start_date')
    )

    groups = [
        ('期限切れ', overdue, 'FFDCE0'),
        ('今週開始予定', starting_soon, 'FFF3CD'),
        ('着手中', in_progress, 'DCE8FF'),
    ]

    headers = [
        'グループ', 'WBS No', 'タスク名', 'ステータス',
        '担当者', '予定開始日', '予定終了日', '実績開始日',
    ]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers))
    _freeze_header(ws)

    for group_label, tasks, fill_hex in groups:
        fill = PatternFill(fill_type='solid', fgColor=fill_hex)
        font = Font(size=9)
        align = Alignment(vertical='center')
        for task in tasks:
            assignees = ', '.join(a.user.username for a in task.assignees.all() if a.user)
            ws.append([
                group_label, task.wbs_no, task.title, task.status,
                assignees,
                _fmt_date(task.start_date), _fmt_date(task.end_date),
                _fmt_date(task.actual_start_date),
            ])
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = align

    _set_col_widths(ws, [14, 10, 40, 12, 18, 14, 14, 14])


# ============================================================
# Sheet 3: 進捗一覧
# ============================================================

def _build_progress_sheet(ws, project):
    all_tasks = _load_tasks(project)
    roots, children_map = _build_children_map(all_tasks)
    nodes = _walk_tree(roots, children_map, visible_ids=None, max_depth=3)
    today = date.today()

    headers = [
        'WBS No', 'タスク名',
        '総時間(h)', 'Done(h)', '進行(h)', '未着手(h)', '進捗%',
        '総件数', 'Done(件)', '進行(件)', '未着手(件)',
        '予定開始日', '予定終了日', '実績開始日', '実績終了日',
        '遅延/巻き(h)',
    ]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers))
    _freeze_header(ws)

    for node in nodes:
        actual = _collect_actual_tasks(node, children_map)
        total_h = sum(float(t.estimated_hours or 0) for t in actual)
        done   = [t for t in actual if t.actual_end_date]
        inprog = [t for t in actual if t.actual_start_date and not t.actual_end_date]
        todo   = [t for t in actual if not t.actual_start_date]

        done_h   = sum(float(t.estimated_hours or 0) for t in done)
        inprog_h = sum(float(t.estimated_hours or 0) for t in inprog)
        todo_h   = sum(float(t.estimated_hours or 0) for t in todo)
        pct = done_h / total_h if total_h > 0 else 0.0

        starts    = [t.start_date for t in actual if t.start_date]
        ends      = [t.end_date for t in actual if t.end_date]
        act_starts = [t.actual_start_date for t in actual if t.actual_start_date]
        act_ends   = [t.actual_end_date for t in actual if t.actual_end_date]

        earliest_start     = min(starts) if starts else None
        latest_end         = max(ends) if ends else None
        earliest_act_start = min(act_starts) if act_starts else None
        latest_act_end     = max(act_ends) if act_ends else None

        is_all_done = len(actual) > 0 and len(done) == len(actual)
        delay_h = 0.0
        if latest_end:
            if latest_end <= today and not is_all_done:
                delay_h = -(total_h - done_h)
            elif latest_end > today and is_all_done:
                delay_h = done_h

        depth = min(node.depth, 3)
        ws.append([
            node.wbs_no,
            _INDENT[depth] + node.title,
            round(total_h, 1), round(done_h, 1), round(inprog_h, 1), round(todo_h, 1),
            f'{pct * 100:.1f}%',
            len(actual), len(done), len(inprog), len(todo),
            _fmt_date(earliest_start), _fmt_date(latest_end),
            _fmt_date(earliest_act_start), _fmt_date(latest_act_end),
            round(delay_h, 1) if delay_h != 0 else '',
        ])

        fill = PatternFill(fill_type='solid', fgColor=_DEPTH_FILLS[depth])
        font = Font(size=9, bold=(node.depth == 0))
        align = Alignment(vertical='center')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = align

        # 遅延列に色を付ける
        delay_cell = ws.cell(row=ws.max_row, column=len(headers))
        if delay_h > 0:
            delay_cell.font = Font(size=9, color='1D4ED8', bold=True)
        elif delay_h < 0:
            delay_cell.font = Font(size=9, color='DC2626', bold=True)

    _set_col_widths(ws, [10, 40, 9, 9, 9, 9, 8, 7, 7, 7, 7, 12, 12, 12, 12, 10])


# ============================================================
# メイン関数
# ============================================================

def export_excel(project, quarter_id=None, start_date=None, end_date=None):
    """Excel ファイルを BytesIO で返す"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'WBS'
    _build_wbs_sheet(ws1, project, quarter_id, start_date, end_date)

    ws2 = wb.create_sheet('直近のタスク')
    _build_recent_sheet(ws2, project, start_date, end_date)

    ws3 = wb.create_sheet('進捗一覧')
    _build_progress_sheet(ws3, project)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
